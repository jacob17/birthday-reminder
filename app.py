import os
import random
import requests
from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError
from dotenv import load_dotenv
from openpyxl import load_workbook
from datetime import datetime, timedelta
import csv
import sys
import json

# Load environment variables from .env
load_dotenv()

# Initialize Slack client
slack_token = os.getenv("SLACK_BOT_TOKEN")
client = WebClient(token=slack_token)

# Set file paths for birthday and coupon files
birthday_file = "birthday_example.xlsx"
coupon_file = "coupon_example.xlsx"

########################################################
# 1) LOAD / SAVE FUNCTIONS
########################################################

def load_birthdays(filename=birthday_file):
    workbook = load_workbook(filename)
    sheet = workbook.active
    birthdays = []
    for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
        birthdays.append({
            "Name": row[0],
            "Slack Display Name": row[1],
            "Slack ID": row[2],
            "Birthday": row[3],
            "Join Date": row[4],
            "Sent": row[5] or "",
            "Locale": row[6] or "en"
        })
    return birthdays

def mark_birthday_as_sent(filename, username, user_id):
    workbook = load_workbook(filename)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_col=7, values_only=False):
        username_cell = row[1]
        slack_id_cell = row[2]
        sent_cell = row[5]
        if slack_id_cell.value == user_id:
            sent_cell.value = "TRUE"
            workbook.save(filename)
            return None
    return None

########################################################
# 2) DATE HELPERS
########################################################

def get_last_sent_birthday_this_year(filename=birthday_file):
    birthdays = load_birthdays(filename)
    sent_dates = []

    today = datetime.now()

    for person in birthdays:
        if str(person["Sent"]).strip().upper() == "TRUE":
            try:
                dt = datetime.strptime(person["Birthday"], "%m/%d/%Y")
                dt_this_year = dt.replace(year=today.year)
                if dt_this_year <= today:
                    sent_dates.append(dt_this_year)
                else:
                    pass
            except ValueError as e:
                continue

    if not sent_dates:
        fallback = today.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        return fallback

    latest = max(sent_dates)
    return latest

def is_birthday_in_dates(birthday_str, dates):
    bday_md = datetime.strptime(birthday_str, "%m/%d/%Y").strftime("%m/%d")
    date_md_list = [d.strftime("%m/%d") for d in dates]

    for d in dates:
        if d.strftime("%m/%d") == bday_md:
            return True
    return False

def is_eligible_birthday_this_year(birthday_str, join_date_str):
    if not birthday_str or not join_date_str:
        return False
    try:
        birthday_date = datetime.strptime(birthday_str, "%m/%d/%Y").replace(year=datetime.now().year)
        join_date = datetime.strptime(join_date_str, "%m/%d/%Y")

        today = datetime.now()
        return birthday_date >= join_date and birthday_date <= today
    except ValueError:
        return False

def get_birthdays_for_dates(dates, filename=birthday_file):
    birthdays = load_birthdays(filename)
    results = []

    for person in birthdays:
        if str(person["Sent"]).strip().upper() != "TRUE":
            if is_birthday_in_dates(person["Birthday"], dates) and is_eligible_birthday_this_year(person["Birthday"], person["Join Date"]):
                results.append(person)

    return results

########################################################
# 3) COUPON & MESSAGE FUNCTIONS
########################################################

def get_coupon_and_mark_sent(filename=coupon_file):
    workbook = load_workbook(filename)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=False):
        link_cell, code_cell, sent_cell = row[0], row[1], row[2]
        if sent_cell.value != "TRUE":
            sent_cell.value = "TRUE"
            workbook.save(filename)
            return link_cell.value, code_cell.value
    return None, None

def is_belated(birthday_str):
    today_md = int(datetime.now().strftime("%m%d"))
    person_md = int(datetime.strptime(birthday_str, "%m/%d/%Y").strftime("%m%d"))
    return person_md < today_md

def is_valid_slack_user(user_id):
    try:
        response = client.users_info(user=user_id)
        return response["ok"]
    except SlackApiError:
        return False
    
def load_translations(path="i18n.json"):
    with open(path, encoding="utf-8") as f:
        return json.load(f)

def send_birthday_message(user_id, username, link, code, birthday_str, locale, translations, custom_text=None):
    belated = is_belated(birthday_str)
    lang = translations.get(locale, translations["en"])
    main_text = custom_text if custom_text else (
        lang["belated_birthday_greeting"] if belated else lang["birthday_greeting"]
    )

    emoji_list = [":happybirthday:", ":meow_birthday:", ":caker_appreciate:", ":cakerloveyou:"]
    random_emoji = random.choice(emoji_list)
    body = lang["message_body"].format(emoji=random_emoji)
    redeem = lang["redeem_link"].format(link=link, code=code)

    message = f"Hi <@{user_id}>, \n{main_text}\n{body}\n{redeem}"

    if not is_valid_slack_user(user_id):
        print(f"[WARNING] Invalid Slack user ID: {user_id}. Skipping message to {username}.")
        invalid_users.append((username, user_id))
        return

    try:
        client.chat_postMessage(channel=user_id, text=message, as_user=True)
        print(f"Message sent successfully to {username} (ID: {user_id}) with link {link} and code {code}.")
    except SlackApiError as e:
        print(f"Error sending message to {username} (ID: {user_id}): {e.response['error']}")

########################################################
# 4) MAIN EXECUTION
########################################################

if __name__ == "__main__":
    invalid_users = []

    print(f"[INFO] Using birthday file: {birthday_file}")
    print(f"[INFO] Using coupon file: {coupon_file}")
    print(f"[INFO] Mode: {'REAL' if 'real' in birthday_file else 'TEST'}")
    translations = load_translations()
    print("""
        Menu:
        1. Send birthday messages
        2. Send coupons to specific person
    """)
    option = input("Enter your choice (1/2): ").strip()

    if option == '2':
        birthdays = load_birthdays(birthday_file)
        selected_people = []

        while True:
            name_input = input("Which user do you want to send coupons to? Enter a Slack Display Name (or type DONE to finish, NONE to abort): ").strip()
            if name_input.lower() == 'none':
                print("Aborted by user.")
                sys.exit(0)
            if name_input.lower() == 'done':
                break
            match = next((p for p in birthdays if p['Slack Display Name'] == name_input), None)
            if match:
                if str(match["Sent"]).strip().upper() == "TRUE":
                    print(f"{match['Slack Display Name']} has already been sent a birthday message. Skipping.")
                else:
                    print(f"Found: {match['Slack Display Name']} (ID: {match['Slack ID']})")
                    selected_people.append(match)
            else:
                print("No match found. Please try again.")

        if selected_people:
            custom_text = input("Enter a custom message for the birthday greeting: ").strip()
            print("Message preview:")
            for person in selected_people:
                locale = person.get("Locale", "en")
                lang = translations.get(locale, translations["en"])
                belated = is_belated(person["Birthday"])
                main_text = custom_text if custom_text else (
                    lang["belated_birthday_greeting"] if belated else lang["birthday_greeting"]
                )
                emoji_list = [":happybirthday:", ":meow_birthday:", ":caker_appreciate:", ":cakerloveyou:"]
                random_emoji = random.choice(emoji_list)
                body = lang["message_body"].format(emoji=random_emoji)
                redeem = lang["redeem_link"].format(link="<LINK>", code="CODE")

                preview = f"Hi @{person['Slack ID']}, \n{main_text}\n{body}\n{redeem}"
                print(preview)
                print("-")

            print(f"The following {len(selected_people)} birthday messages will be sent:")
            for person in selected_people:
                print(f" - {person['Slack Display Name']} ({person['Birthday']})")

            for person in selected_people:
                coupon_link, coupon_code = get_coupon_and_mark_sent(coupon_file)
                if coupon_link and coupon_code:
                    send_birthday_message(
                        user_id=person["Slack ID"],
                        username=person["Slack Display Name"],
                        link=coupon_link,
                        code=coupon_code,
                        birthday_str="01/01/2000",  # Placeholder date; only used if custom_text is not provided
                        locale=person["Locale"],
                        translations=translations,
                        custom_text=custom_text
                    )
                else:
                    print("No unused coupons available.")
        else:
            print("No valid users entered.")
        sys.exit(0)

    elif option != '1':
        print("Invalid option. Exiting.")
        sys.exit(1)
        
    today = datetime.now()
    last_sent_date = get_last_sent_birthday_this_year(birthday_file)

    dates_to_check = [last_sent_date + timedelta(days=i) for i in range(0, (today - last_sent_date).days + 1)]
    if today not in dates_to_check:
        dates_to_check.append(today)

    birthday_people = get_birthdays_for_dates(dates_to_check, birthday_file)

    if birthday_people:
        print(f"The following {len(birthday_people)} birthday messages will be sent:")
        for person in birthday_people:
            print(f" - {person['Slack Display Name']} ({person['Birthday']})")

        confirm = input("Do you want to send these messages? (Y/N): ").strip().lower()
        if confirm != 'y':
            print("Aborted by user.")
            sys.exit(0)

        print(f"Found {len(birthday_people)} birthdays.")

        for person in birthday_people:
            coupon_link, coupon_code = get_coupon_and_mark_sent(coupon_file)
            if coupon_link and coupon_code:
                send_birthday_message(
                    user_id=person["Slack ID"],
                    username=person["Slack Display Name"],
                    link=coupon_link,
                    code=coupon_code,
                    birthday_str=person["Birthday"],
                    locale=person["Locale"],
                    translations=translations
                )
                mark_birthday_as_sent(birthday_file, person["Slack Display Name"], person["Slack ID"])
            else:
                print("No unused coupons available.")
    else:
        print("No birthdays found for the specified dates.")

    if invalid_users:
        print("The following Slack IDs were invalid and skipped:")
        for name, uid in invalid_users:
            print(f" - {uid} ({name})")
